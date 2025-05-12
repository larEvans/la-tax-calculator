import sys, os
from datetime import datetime,date
from flask import Flask, render_template_string, request, send_file, redirect, url_for, flash, session
from flask_sqlalchemy import SQLAlchemy
import pandas as pd
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference
from collections import defaultdict
if getattr(sys, 'frozen', False):
    basedir = os.path.dirname(sys.executable)
else:
    basedir = os.path.abspath(os.path.dirname(__file__))

db_path   = os.path.join(basedir, 'entries.db')
saved_dir = os.path.join(basedir, 'saved_entries')
os.makedirs(saved_dir, exist_ok=True)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'replace_with_real_secret'
app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{db_path}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

class Entry(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    title      = db.Column(db.String(255), nullable=False)   # <-- new!
    timestamp  = db.Column(db.DateTime, default=datetime.utcnow)
    tax_csv    = db.Column(db.Text, nullable=False)
    exp_csv    = db.Column(db.Text, nullable=False)
    final_csv  = db.Column(db.Text, nullable=False)

    incomes    = db.relationship(
        'Income',
        backref='entry',
        cascade='all, delete-orphan'
    )

class Income(db.Model):
    __tablename__ = 'income'
    id          = db.Column(db.Integer, primary_key=True)
    entry_id    = db.Column(
        db.Integer,
        db.ForeignKey('entry.id', ondelete='CASCADE'),
        nullable=False
    )
    sender      = db.Column(db.String(80),  nullable=False)
    Gross       = db.Column(db.Float,      nullable=False)
    income_type = db.Column(db.String(20), nullable=False)
    date        = db.Column(db.Date,       default=datetime.utcnow().date)


with app.app_context():
    db.create_all()

federal_brackets = [
    (0, 11000, 0.10),
    (11000, 44725, 0.12),
    (44725, 95375, 0.22),
    (95375, 182100,0.24),
    (182100,231250,0.32),
    (231250,578125,0.35),
    (578125,float('inf'),0.37),
]
louisiana_tax_rate = 0.04

def calculate_federal_tax(income: float) -> float:
    tax = 0.0
    for lo, hi, rate in federal_brackets:
        if income > lo:
            taxable = min(income, hi) - lo
            tax += taxable * rate
        else:
            break
    return tax


base_style = '''
<style>
  body { 
    font-family: Arial, sans-serif;
    padding: 20px;
    background: #f4f7f9; color: #333; 
    display: flex;
    flex-direction:column;
    align-items: center;
    min-height: 100vh; 
  }


  fieldset {
    background: #fff;
    padding: 10px;
    border: 1px solid #ccc;
    margin-bottom: 15px;
    border-radius: 4px;
  }

  @media (prefers-color-scheme: dark) {
    body { background: #121212; color: #e0e0e0; }
    table, th, td { border-color: #444; }
    input, button, .btn-link { background: #1f1f1f; color: #e0e0e0; border-color: #444; }
    /* dark‐mode fieldsets now match inputs/buttons */
    fieldset {
      background: #1f1f1f !important;
      border-color: #444;
    }
    nav a { color: #bb86fc; }
  }

  nav a, .btn-link, button {
    background: #006d77;
    color: #fff;
    padding: 8px 12px;
    margin: 5px;
    border: none;
    border-radius: 4px;
    text-decoration: none;
    cursor: pointer;
  }
  nav a:hover, button:hover, .btn-link:hover { background: #005f68; }

  table { width: 100%; border-collapse: collapse; margin-bottom: 20px; }
  th, td { border: 1px solid #ccc; padding: 8px; text-align: left; }
  th { background: #006d77; color: #fff; }

  label { display: block; margin: 5px 0; }
  input { width: calc(100% - 14px); padding: 6px; border: 1px solid #ccc; border-radius: 4px; }
    .chart-container {
    display: flex;
    justify-content: space-around;
    align-items: flex-start;
    gap: 20px;
    margin-bottom: 20px;
  }
  .chart-block {
    text-align: center;
  }
</style>
'''
nav_html = '<nav><a href="/">Home</a> | <a href="/saved-entries">Saved Entries</a> | <a href="/statements">Statements</a></nav>'

index_html = base_style + '''
<html>
  <head><title>Step 1: Checks</title></head>
  <body>
''' + nav_html + '''
    <h2>Louisana Tax and Budget Calculator</h2>
    <h3>Enter the number of checks your entering:</h3>
    <form method="post" action="/tax-entry">
      <label>Checks: <input name="num_checks" type="number" min="1" value="1" required></label>
      <button type="submit">Next</button>
    </form>
  </body>
</html>
'''

tax_entry_html = base_style + '''
<html>
  <head><title>Step 2: Gross</title></head>
  <body>
''' + nav_html + '''
    <h2>{% if edit %}Edit{% else %}Enter{% endif %} Gross Amounts</h2>
    <form method="post" action="/show-taxes">
      <input type="hidden" name="num_checks" value="{{ n }}">
      {% if edit %}
        <input type="hidden" name="tax_csv" value="{{ tax_csv }}">
      {% endif %}
      {% for i in range(n) %}
      <fieldset>
        <legend>Check {{ i+1 }}</legend>
        <label>Sender:
          <input name="sender_{{i}}">
        </label>
        <label>Gross:
          <input name="Gross_{{i}}" type="number" step="0.01">
        </label>
        <label>Type:
          <select name="type_{{i}}">
            <option value="1099-NEC">1099-NEC</option>
            <option value="W-2">W-2</option>
            <option value="Retirement">Retirement</option>
          </select>
        </label>
        <label>Date:
          <input name="date_{{i}}" type="date" value="{{ today }}" />
        </label>
      </fieldset>

      {% endfor %}
      <button type="submit">Calculate Taxes</button>
    </form>
  </body>
</html>
'''

show_taxes_html = base_style + '''
<html>
  <head>
    <title>Step 3: Tax Breakdown</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
  </head>
  <body>
''' + nav_html + '''
    <h2>Tax Breakdown</h2>
    <!-- inside show_taxes_html -->
    <table>
      <tr>
        {% for c in cols %}<th>{{ c }}</th>{% endfor %}
      </tr>
      {% for row in rows %}
        <tr>
          {% for v in row %}<td>{{ v }}</td>{% endfor %}
        </tr>
      {% endfor %}
    </table>


    <h3>Components</h3>
    <canvas id="taxChart" width="200" height="200" style="margin:auto;display:block;"></canvas>
    <form method="post" action="/expense-entry">
      <input type="hidden" name="tax_csv"  value="{{ csv }}">
      <input type="hidden" name="exp_csv"  value="{{ exp_csv }}">   <!-- NEW -->
      <input type="hidden" name="num_checks" value="{{ rows|length }}">
      <button type="submit">Next</button>
    </form>
    <script>
      new Chart(document.getElementById('taxChart'), {
        type: 'pie',
        data: { labels: {{ comp_labels|tojson }}, datasets: [{ data: {{ comp_data|tojson }} }] },
        options: { responsive: false }
      });
    </script>
  </body>
</html>
'''
expense_entry_html = base_style + '''
<html>
  <head>
    <title>Step 4: Expenses</title>
    <style>
      /* keep your white fieldset in light mode and dark in dark-mode */
      fieldset { background:#fff; padding:10px; border:1px solid #ccc; border-radius:4px; margin-bottom:15px; }
      @media (prefers-color-scheme: dark) {
        fieldset { background:#1f1f1f; border-color:#444; }
      }
    </style>
  </head>
  <body>
''' + nav_html + '''
    <h2>Enter Expenses</h2>
    <form method="post" action="/show-final" id="expForm">
      <input type="hidden" name="tax_csv" value="{{ tax_csv }}">
    {% set saved = session.get('exp_data', {}) %}
  {% for i in range(num_checks) %}
    {% set cnt = (saved.get('count_' ~ i|string) or 1) | int %}
    <fieldset>
      <legend>Check {{ i+1 }}</legend>
      <input type="hidden" id="count_{{ i }}" name="count_{{ i }}" value="{{ cnt }}">
      <div id="expenses_{{ i }}">
        {% for j in range(cnt) %}
          <label>Name:
            <input name="exp_name_{{ i }}_{{ j }}"
                  value="{{ saved.get('exp_name_' ~ i ~ '_' ~ j, '') }}"
                  required>
          </label>
          <label>Amount:
            <input name="exp_amt_{{ i }}_{{ j }}" type="number" step="0.01"
                  value="{{ saved.get('exp_amt_' ~ i ~ '_' ~ j, '') }}"
                  required>
          </label>
          <button type="button" onclick="removeExpense(this)">Remove</button>
        {% endfor %}
    </div>
    <button type="button" onclick="addExpense({{ i }})">Add Expense</button>
  </fieldset>
{% endfor %}

      <button type="submit">Finish</button>
    </form>

    <script>
    // Restore any saved entries from localStorage
    document.addEventListener('DOMContentLoaded', () => {
      const saved = JSON.parse(localStorage.getItem('exp_data')||'{}');
      for (let key in saved) {
        const el = document.querySelector(`[name="${key}"]`);
        if (el) el.value = saved[key];
      }
      // fix counters
      Object.keys(saved).forEach(k => {
        if (k.startsWith('count_')) {
          const i = k.split('_')[1];
          document.getElementById('count_'+i).value = saved[k];
        }
      });
    });

    // Before leaving the page, write all expense inputs & counts into localStorage
    window.addEventListener('beforeunload', () => {
      const data = {};
      document.querySelectorAll('#expForm [name]').forEach(el => {
        if (el.name.startsWith('exp_name_') ||
            el.name.startsWith('exp_amt_')  ||
            el.name.startsWith('count_')) {
          data[el.name] = el.value;
        }
      });
      localStorage.setItem('exp_data', JSON.stringify(data));
    });

     function addExpense(i) {
        const cont = document.getElementById('expenses_' + i);
        const countInput = document.getElementById('count_' + i);
        let cnt = parseInt(countInput.value, 10);
        const rowHtml = `
          <div class="expense-row" data-check="${i}">
            <label>Name:   <input name="exp_name_${i}_${cnt}" required></label>
            <label>Amount: <input name="exp_amt_${i}_${cnt}" type="number" step="0.01" required></label>
            <button type="button" onclick="removeExpense(this)">Remove</button>
          </div>`;
        cont.insertAdjacentHTML('beforeend', rowHtml);
        countInput.value = ++cnt;
      }

      function removeExpense(btn) {
        const row = btn.closest('.expense-row');
        const i = row.dataset.check;
        row.remove();

        // update the hidden counter
        const newCount = document.querySelectorAll(
          `#expenses_${i} .expense-row`
        ).length;
        document.getElementById('count_' + i).value = newCount;
      }
    </script>
  </body>
</html>
'''



final_html = base_style + '''
<html>
  <head>
    <title>Step 5: Summary</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
  </head>
  <body>
''' + nav_html + '''
    <h2>Final Summary (Editable)</h2>
    <form id="finalForm">
      <h3>Tax Breakdown</h3>
      <table>
        <tr>{% for c in tax_cols %}<th>{{ c }}</th>{% endfor %}</tr>
        {% for row in tax_rows %}
          <tr>{% for v in row %}<td>{{ v }}</td>{% endfor %}</tr>
        {% endfor %}
      </table>

      <h3>Expenses & Net Profit</h3>
      <table id="expTable">
        <thead>
          <tr>
            <th>Sender</th>
            <th>Expense</th>
            <th>Amount</th>
            <th>Net After</th>
            <th>Action</th>
          </tr>
        </thead>
        <tbody>
          {% for s,e,a,net_after in exp_rows %}
            <tr data-sender="{{ s }}">
              <td>{{ s }}</td>
              <td><input name="exp_name_{{ loop.index0 }}" value="{{ e }}"></td>
              <td><input name="exp_amt_{{ loop.index0 }}" value="{{ a }}" oninput="recalculate()"></td>
              <td><input name="net_after_{{ loop.index0 }}" value="{{ net_after }}" readonly></td>
              <td><button type="button" onclick="removeRow(this)">Remove</button></td>
            </tr>
          {% endfor %}
        </tbody>
        <tfoot>
          <tr>
            <th colspan="3">Final Net</th>
            <th colspan="2" id="finalNetCell"><strong>${{ total_net }}</strong></th>
          </tr>
        </tfoot>
      </table>

      {% if not view_only %}
        <button type="button" onclick="window.history.back()">Back</button>
      {% endif %}      
      <button type="button" onclick="addRow()">Add Expense</button>
      <button type="button" onclick="recalculate()">Recalculate</button>
      <button type="button" onclick="saveEntry()">Save Entry</button>
      <button formaction="/download-final" formmethod="post">Export to Excel</button>

      <input type="hidden" name="tax_csv"   value="{{ tax_csv }}">
      <input type="hidden" name="exp_csv"   value="{{ exp_csv }}">
      <input type="hidden" name="final_csv" value="{{ final_csv }}">
    </form>

    <div class="chart-container">
  <div class="chart-block">
    <h3>Tax Breakdown Pie</h3>
    <canvas id="taxBreakdownChart" width="200" height="200"></canvas>
  </div>
  <div class="chart-block">
    <h3>Overall Distribution</h3>
    <canvas id="finalChart" width="200" height="200"></canvas>
  </div>
</div>
    <script>
      const origNet = {{ orig_nets|tojson }};

      function recalculate() {
        // Sequentially subtract each expense from its sender’s net
        const running = {...origNet};
        document.querySelectorAll('#expTable tbody tr[data-sender]').forEach(row => {
          const sender = row.dataset.sender;
          const idx    = row.querySelector('input[name^="exp_amt_"]').name.split('_').pop();
          const amt    = parseFloat(row.querySelector(`input[name="exp_amt_${idx}"]`).value) || 0;
          running[sender] -= amt;
          row.querySelector(`input[name="net_after_${idx}"]`).value = running[sender].toFixed(2);
        });

        // Compute overall final net
        let overall = Object.values(running).reduce((sum,v) => sum + v, 0);
        document.getElementById('finalNetCell').innerHTML = `<strong>$${overall.toFixed(2)}</strong>`;
      }

      document.querySelectorAll('#expTable input[name^="exp_amt_"]')
        .forEach(input => input.addEventListener('input', recalculate));
      document.addEventListener('DOMContentLoaded', recalculate);

      function removeRow(btn) { btn.closest('tr').remove(); recalculate(); }

      function addRow() {
        const tbody = document.querySelector('#expTable tbody');
        const idx = tbody.rows.length;
        const row = document.createElement('tr');
        row.setAttribute('data-sender', '');
        row.innerHTML = `
          <td><input name="exp_sender_${idx}" placeholder="Sender"
                    onchange="this.closest('tr').setAttribute('data-sender', this.value)"></td>
          <td><input name="exp_name_${idx}" placeholder="Expense"></td>
          <td><input name="exp_amt_${idx}" oninput="recalculate()"></td>
          <td><input name="net_after_${idx}" readonly></td>
          <td><button type="button" onclick="removeRow(this)">Remove</button></td>
        `;
        tbody.appendChild(row);
        // immediately recalc so new row's net shows up
        recalculate();
      }

    async function saveEntry() {
        // ask the user for a name
        const title = prompt("Enter a name for this entry:");
        if (!title) return;  // user cancelled

        const form = document.getElementById('finalForm');
        const data = new FormData(form);
        data.append('title', title);

        const resp = await fetch('/save-entry', {
          method: 'POST',
          body: data,
          headers: { 'X-Requested-With': 'XMLHttpRequest' }
        });

        if (resp.ok || resp.status===302) {
          alert('Entry "' + title + '" saved successfully!');
          // optionally redirect back to Saved Entries:
          window.location = '/saved-entries';
        } else {
          alert('Save failed.');
        }
      }
      new Chart(
        document.getElementById('taxBreakdownChart'),
        { type:'pie',
          data:{ labels: {{ comp_labels|tojson }},
                 datasets:[{ data: {{ comp_data|tojson }} }]},
          options:{ responsive:false }
        }
      );
      new Chart(
        document.getElementById('finalChart'),
        { type:'pie',
          data:{ labels:['Tax','Expenses','Net'],
                 datasets:[{ data:[{{ total_tax }},{{ total_exp }},{{ total_net }}] }]},
          options:{ responsive:false }
        }
      );
    </script>
  </body>
</html>
'''
statements_html = base_style + nav_html + '''
<html>
  <head><title>Income Statements</title></head>
  <body>
    <h2>Income & Expense Statements by Month</h2>

    <form method="get" style="margin-bottom:1em;">
      <label>Filter by type:
        <select name="type" onchange="this.form.submit()">
          {% for t in types %}
            <option value="{{t}}" {% if t==filter_type %}selected{% endif %}>{{t}}</option>
          {% endfor %}
        </select>
      </label>
    </form>
    <form style="margin-bottom:1em;" action="{{ url_for('download_statements') }}" method="get">
      <input type="hidden" name="type" value="{{ filter_type }}">
      <button type="submit">Export Statements to Excel</button>
    </form>

    {% if not summary %}
      <p><em>No data to show.</em></p>
    {% else %}
      {% for row in summary %}
        <details open style="margin-bottom:2em;">
          <summary style="font-size:1.1em; cursor:pointer;">
            {{ row.month.strftime('%Y-%m') }}
            — Income: ${{ '%.2f'|format(row.inc) }}
            | Expenses: ${{ '%.2f'|format(row.exp) }}
            | Taxes Due: ${{ '%.2f'|format(row.tax) }}
          </summary>
          <div style="padding: 0.5em 1em;">
            <h4>Incomes this month</h4>
            <table>
              <tr>
                <th>Date</th><th>Sender</th><th>Type</th><th>Gross</th><th>Taxes</th><th>Entry</th>
              </tr>
              {% for it in inc_groups[row.month] %}
                <tr>
                  <td>{{ it.date.strftime('%Y-%m-%d') }}</td>
                  <td>{{ it.sender }}</td>
                  <td>{{ it.type }}</td>
                  <td>${{ '%.2f'|format(it.Gross) }}</td>
                  <td>${{ '%.2f'|format(it.taxes_due) }}</td>
                  <td><a href="{{ url_for('view_entry', entry_id=it.entry_id) }}">View</a></td>
                </tr>
              {% endfor %}
            </table>

            <h4 style="margin-top:1em;">Expenses this month</h4>
<table>
  <tr><th>Date</th><th>Sender</th><th>Expense</th><th>Amount</th><th>Entry</th></tr>
  {% for ex in exp_groups[row.month] %}
  <tr>
    <!-- now use ex.date -->
    <td>{{ ex.date.strftime('%Y-%m-%d') }}</td>
    <td>{{ ex.sender }}</td>
    <td>{{ ex.name }}</td>
    <td>${{ '%.2f'|format(ex.amt) }}</td>
    <td><a href="{{ url_for('view_entry', entry_id=ex.entry_id) }}">View</a></td>
  </tr>
  {% endfor %}
</table>
          </div>
        </details>
      {% endfor %}

      <h2>Yearly Summary</h2>
      <table>
        <tr><th>Year</th><th>Total Gross</th><th>Total Expenses</th><th>Total Taxes</th></tr>
        {% for y in yearly %}
          <tr>
            <td>{{ y.year }}</td>
            <td>${{ '%.2f'|format(y.inc_total) }}</td>
            <td>${{ '%.2f'|format(y.exp_total) }}</td>
            <td>${{ '%.2f'|format(y.tax_total) }}</td>
          </tr>
        {% endfor %}
      </table>
    {% endif %}
  </body>
</html>
'''

 
def se_tax(amount: float) -> float:
    return amount * 0.153

def fed_tax(amount: float) -> float:
    return calculate_federal_tax(amount)

def state_tax(amount: float) -> float:
    return amount * louisiana_tax_rate

def show_final_context(tax_csv, exp_csv, final_csv):
    df_tax = pd.read_csv(StringIO(tax_csv))
    comp_labels = ['Self-EE','Fed','State']
    comp_data = [
        df_tax['Self-EE Tax'].sum(),
        df_tax['Fed Tax'].sum(),
        df_tax['State Tax'].sum()
    ]
    exp_lines = []
    total_exp = 0.0
    for i, sender in enumerate(df_tax['Sender']):
        count = int(request.form.get(f'count_{i}', 0))
        for j in range(count):
            name = request.form[f'exp_name_{i}_{j}']
            amt = float(request.form[f'exp_amt_{i}_{j}'])
            exp_lines.append((sender, name, amt))
            total_exp += amt
    total_tax = df_tax['Total Tax'].sum()
    total_net = df_tax['Net'].sum() - total_exp
    orig_nets = df_tax.set_index('Sender')['Net'].to_dict()
    exp_rows = []
    for sender, name, amt in exp_lines:
        total_sender_exp = sum(amt for s, _, amt in exp_lines if s == sender)
        net_profit = round(orig_nets[sender] - total_sender_exp, 2)
        exp_rows.append([sender, name, amt, net_profit])
    tax_rows = [row[1:] for row in df_tax.values.tolist()]
    tax_cols = df_tax.columns.tolist()[1:]
    return dict(
        tax_cols=tax_cols,
        tax_rows=tax_rows,
        exp_rows=exp_rows,
        total_tax=round(total_tax,2),
        total_exp=round(total_exp,2),
        total_net=round(total_net,2),
        comp_labels=comp_labels,
        comp_data=comp_data,
        tax_csv=tax_csv,
        exp_csv=exp_csv,
        final_csv=final_csv,
        orig_nets=orig_nets
    )

@app.route('/', methods=['GET'])
def index():
    return render_template_string(index_html)

@app.route('/tax-entry', methods=['POST'])
def tax_entry():
    n     = int(request.form['num_checks'])
    today = date.today().isoformat()   # "2025-05-08"
    return render_template_string(
      tax_entry_html,
      n=n,
      edit=False,
      senders=[],
      grosses=[],
      today=today
    )
from datetime import datetime
import pandas as pd

@app.route('/show-taxes', methods=['POST'])
def show_taxes():
    n = int(request.form['num_checks'])
    records = []
    sums = {'se': 0.0, 'fed': 0.0, 'st': 0.0}

    for i in range(n):
        sender   = request.form[f'sender_{i}']
        Gross    = float(request.form[f'Gross_{i}'])
        inc_type = request.form[f'type_{i}']

        date_str   = request.form[f'date_{i}']
        check_date = datetime.fromisoformat(date_str).date()

        if inc_type == '1099-NEC':
            se  = Gross * 0.153
            fed = calculate_federal_tax(Gross)
            st  = Gross * louisiana_tax_rate
        else:
            se = fed = st = 0.0

        total_tax = se + fed + st
        net       = Gross - total_tax

        sums['se']  += se
        sums['fed'] += fed
        sums['st']  += st

        records.append([
            sender,
            inc_type,
            check_date,
            round(Gross,    2),
            round(se,       2),
            round(fed,      2),
            round(st,       2),
            round(total_tax,2),
            round(net,      2)
        ])

    df = pd.DataFrame(records, columns=[
        'Sender','Type','Date',
        'Gross','Self-EE Tax','Fed Tax','State Tax','Total Tax','Net'
    ])


import sys, os
from datetime import datetime,date
from flask import Flask, render_template_string, request, send_file, redirect, url_for, flash, session
from flask_sqlalchemy import SQLAlchemy
import pandas as pd
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference
from collections import defaultdict
if getattr(sys, 'frozen', False):
    basedir = os.path.dirname(sys.executable)
else:
    basedir = os.path.abspath(os.path.dirname(__file__))

db_path   = os.path.join(basedir, 'entries.db')
saved_dir = os.path.join(basedir, 'saved_entries')
os.makedirs(saved_dir, exist_ok=True)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'replace_with_real_secret'
app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{db_path}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

class Entry(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    title      = db.Column(db.String(255), nullable=False)   # <-- new!
    timestamp  = db.Column(db.DateTime, default=datetime.utcnow)
    tax_csv    = db.Column(db.Text, nullable=False)
    exp_csv    = db.Column(db.Text, nullable=False)
    final_csv  = db.Column(db.Text, nullable=False)

    incomes    = db.relationship(
        'Income',
        backref='entry',
        cascade='all, delete-orphan'
    )

class Income(db.Model):
    __tablename__ = 'income'
    id          = db.Column(db.Integer, primary_key=True)
    entry_id    = db.Column(
        db.Integer,
        db.ForeignKey('entry.id', ondelete='CASCADE'),
        nullable=False
    )
    sender      = db.Column(db.String(80),  nullable=False)
    Gross       = db.Column(db.Float,      nullable=False)
    income_type = db.Column(db.String(20), nullable=False)
    date        = db.Column(db.Date,       default=datetime.utcnow().date)


with app.app_context():
    db.create_all()

federal_brackets = [
    (0, 11000, 0.10),
    (11000, 44725, 0.12),
    (44725, 95375, 0.22),
    (95375, 182100,0.24),
    (182100,231250,0.32),
    (231250,578125,0.35),
    (578125,float('inf'),0.37),
]
louisiana_tax_rate = 0.04

def calculate_federal_tax(income: float) -> float:
    tax = 0.0
    for lo, hi, rate in federal_brackets:
        if income > lo:
            taxable = min(income, hi) - lo
            tax += taxable * rate
        else:
            break
    return tax


base_style = '''
<style>
  body { 
    font-family: Arial, sans-serif;
    padding: 20px;
    background: #f4f7f9; color: #333; 
    display: flex;
    flex-direction:column;
    align-items: center;
    min-height: 100vh; 
  }


  fieldset {
    background: #fff;
    padding: 10px;
    border: 1px solid #ccc;
    margin-bottom: 15px;
    border-radius: 4px;
  }

  @media (prefers-color-scheme: dark) {
    body { background: #121212; color: #e0e0e0; }
    table, th, td { border-color: #444; }
    input, button, .btn-link { background: #1f1f1f; color: #e0e0e0; border-color: #444; }
    /* dark‐mode fieldsets now match inputs/buttons */
    fieldset {
      background: #1f1f1f !important;
      border-color: #444;
    }
    nav a { color: #bb86fc; }
  }

  nav a, .btn-link, button {
    background: #006d77;
    color: #fff;
    padding: 8px 12px;
    margin: 5px;
    border: none;
    border-radius: 4px;
    text-decoration: none;
    cursor: pointer;
  }
  nav a:hover, button:hover, .btn-link:hover { background: #005f68; }

  table { width: 100%; border-collapse: collapse; margin-bottom: 20px; }
  th, td { border: 1px solid #ccc; padding: 8px; text-align: left; }
  th { background: #006d77; color: #fff; }

  label { display: block; margin: 5px 0; }
  input { width: calc(100% - 14px); padding: 6px; border: 1px solid #ccc; border-radius: 4px; }
    .chart-container {
    display: flex;
    justify-content: space-around;
    align-items: flex-start;
    gap: 20px;
    margin-bottom: 20px;
  }
  .chart-block {
    text-align: center;
  }
</style>
'''
nav_html = '<nav><a href="/">Home</a> | <a href="/saved-entries">Saved Entries</a> | <a href="/statements">Statements</a></nav>'

index_html = base_style + '''
<html>
  <head><title>Step 1: Checks</title></head>
  <body>
''' + nav_html + '''
    <h2>Louisana Tax and Budget Calculator</h2>
    <h3>Enter the number of checks your entering:</h3>
    <form method="post" action="/tax-entry">
      <label>Checks: <input name="num_checks" type="number" min="1" value="1" required></label>
      <button type="submit">Next</button>
    </form>
  </body>
</html>
'''

tax_entry_html = base_style + '''
<html>
  <head><title>Step 2: Gross</title></head>
  <body>
''' + nav_html + '''
    <h2>{% if edit %}Edit{% else %}Enter{% endif %} Gross Amounts</h2>
    <form method="post" action="/show-taxes">
      <input type="hidden" name="num_checks" value="{{ n }}">
      {% if edit %}
        <input type="hidden" name="tax_csv" value="{{ tax_csv }}">
      {% endif %}
      {% for i in range(n) %}
      <fieldset>
        <legend>Check {{ i+1 }}</legend>
        <label>Sender:
          <input name="sender_{{i}}" required>
        </label>
        <label>Gross:
          <input name="Gross_{{i}}" type="number" step="0.01" required>
        </label>
        <label>Type:
          <select name="type_{{i}}">
            <option value="1099-NEC">1099-NEC</option>
            <option value="W-2">W-2</option>
            <option value="Retirement">Retirement</option>
          </select>
        </label>
        <label>Date:
          <input name="date_{{i}}" type="date" value="{{ today }}" />
        </label>
      </fieldset>

      {% endfor %}
      <button type="submit">Calculate Taxes</button>
    </form>
  </body>
</html>
'''

show_taxes_html = base_style + '''
<html>
  <head>
    <title>Step 3: Tax Breakdown</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
  </head>
  <body>
''' + nav_html + '''
    <h2>Tax Breakdown</h2>
    <!-- inside show_taxes_html -->
    <table>
      <tr>
        {% for c in cols %}<th>{{ c }}</th>{% endfor %}
      </tr>
      {% for row in rows %}
        <tr>
          {% for v in row %}<td>{{ v }}</td>{% endfor %}
        </tr>
      {% endfor %}
    </table>


    <h3>Components</h3>
    <canvas id="taxChart" width="200" height="200" style="margin:auto;display:block;"></canvas>
    <form method="post" action="/expense-entry">
      <input type="hidden" name="tax_csv"  value="{{ csv }}">
      <input type="hidden" name="exp_csv"  value="{{ exp_csv }}">   <!-- NEW -->
      <input type="hidden" name="num_checks" value="{{ rows|length }}">
      <button type="submit">Next</button>
    </form>
    <script>
      new Chart(document.getElementById('taxChart'), {
        type: 'pie',
        data: { labels: {{ comp_labels|tojson }}, datasets: [{ data: {{ comp_data|tojson }} }] },
        options: { responsive: false }
      });
    </script>
  </body>
</html>
'''
expense_entry_html = base_style + '''
<html>
  <head>
    <title>Step 4: Expenses</title>
    <style>
      /* keep your white fieldset in light mode and dark in dark-mode */
      fieldset { background:#fff; padding:10px; border:1px solid #ccc; border-radius:4px; margin-bottom:15px; }
      @media (prefers-color-scheme: dark) {
        fieldset { background:#1f1f1f; border-color:#444; }
      }
    </style>
  </head>
  <body>
''' + nav_html + '''
    <h2>Enter Expenses</h2>
    <p>Enter your expenses now or later. You can always come back and add or modify your expenses.</p>
    <form method="post" action="/show-final" id="expForm">
      <input type="hidden" name="tax_csv" value="{{ tax_csv }}">
    {% set saved = session.get('exp_data', {}) %}
  {% for i in range(num_checks) %}
    {% set cnt = (saved.get('count_' ~ i|string) or 1) | int %}
    <fieldset>
      <legend>Expenses for step 1: {{ i+1 }}</legend>
      <input type="hidden" id="count_{{ i }}" name="count_{{ i }}" value="{{ cnt }}">
      <div id="expenses_{{ i }}">
  {% for j in range(cnt) %}
    <div class="expense-row" data-check="{{ i }}">
      <label>Name:
        <input name="exp_name_{{ i }}_{{ j }}"
              value="{{ saved.get('exp_name_' ~ i ~ '_' ~ j, '') }}">
      </label>
      <label>Amount:
        <input name="exp_amt_{{ i }}_{{ j }}" type="number" step="0.01"
              value="{{ saved.get('exp_amt_' ~ i ~ '_' ~ j, '') }}">
      </label>
      <button type="button" onclick="removeExpense(this)">Remove</button>
    </div>
  {% endfor %}
</div>

    </div>
    <button type="button" onclick="addExpense({{ i }})">Add Expense</button>
  </fieldset>
{% endfor %}

      <button type="submit">Finish</button>
    </form>

    <script>
    // Restore any saved entries from localStorage
    document.addEventListener('DOMContentLoaded', () => {
      const saved = JSON.parse(localStorage.getItem('exp_data')||'{}');
      for (let key in saved) {
        const el = document.querySelector(`[name="${key}"]`);
        if (el) el.value = saved[key];
      }
      // fix counters
      Object.keys(saved).forEach(k => {
        if (k.startsWith('count_')) {
          const i = k.split('_')[1];
          document.getElementById('count_'+i).value = saved[k];
        }
      });
    });

    // Before leaving the page, write all expense inputs & counts into localStorage
    window.addEventListener('beforeunload', () => {
      const data = {};
      document.querySelectorAll('#expForm [name]').forEach(el => {
        if (el.name.startsWith('exp_name_') ||
            el.name.startsWith('exp_amt_')  ||
            el.name.startsWith('count_')) {
          data[el.name] = el.value;
        }
      });
      localStorage.setItem('exp_data', JSON.stringify(data));
    });

     function addExpense(i) {
  const cont = document.getElementById('expenses_' + i);
  const countInput = document.getElementById('count_' + i);
  let cnt = parseInt(countInput.value, 10);
  const rowHtml = `
    <div class="expense-row" data-check="${i}">
      <label>Name:   <input name="exp_name_${i}_${cnt}"></label>
      <label>Amount: <input name="exp_amt_${i}_${cnt}" type="number" step="0.01"></label>
      <button type="button" onclick="removeExpense(this)">Remove</button>
    </div>`;
  cont.insertAdjacentHTML('beforeend', rowHtml);
  countInput.value = ++cnt;
}

      function removeExpense(btn) {
  const row = btn.closest('.expense-row');
  const i   = row.dataset.check;
  row.remove();
  // update the counter so that show-final skips entirely if zero
  const newCount = document.querySelectorAll(
    `#expenses_${i} .expense-row`
  ).length;
  document.getElementById('count_' + i).value = newCount;
}
    </script>
  </body>
</html>
'''



final_html = base_style + '''
<html>
  <head>
    <title>Step 5: Summary</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
  </head>
  <body>
''' + nav_html + '''
    <h2>Final Summary (Editable)</h2>
    <form id="finalForm">
      <h3>Tax Breakdown</h3>
      <table>
        <tr>{% for c in tax_cols %}<th>{{ c }}</th>{% endfor %}</tr>
        {% for row in tax_rows %}
          <tr>{% for v in row %}<td>{{ v }}</td>{% endfor %}</tr>
        {% endfor %}
      </table>

      <h3>Expenses & Net Profit</h3>
      <table id="expTable">
        <thead>
          <tr>
            <th>Sender</th>
            <th>Expense</th>
            <th>Amount</th>
            <th>Net After</th>
            <th>Action</th>
          </tr>
        </thead>
        <tbody>
          {% for s,e,a,net_after in exp_rows %}
            <tr data-sender="{{ s }}">
              <td>{{ s }}</td>
              <td><input name="exp_name_{{ loop.index0 }}" value="{{ e }}"></td>
              <td><input name="exp_amt_{{ loop.index0 }}" value="{{ a }}" oninput="recalculate()"></td>
              <td><input name="net_after_{{ loop.index0 }}" value="{{ net_after }}" readonly></td>
              <td><button type="button" onclick="removeRow(this)">Remove</button></td>
            </tr>
          {% endfor %}
        </tbody>
        <tfoot>
          <tr>
            <th colspan="3">Final Net</th>
            <th colspan="2" id="finalNetCell"><strong>${{ total_net }}</strong></th>
          </tr>
        </tfoot>
      </table>

      {% if not view_only %}
        <button type="button" onclick="window.history.back()">Back</button>
      {% endif %}      
      <button type="button" onclick="addRow()">Add Expense</button>
      <button type="button" onclick="recalculate()">Recalculate</button>
      <button type="button" onclick="saveEntry()">Save Entry</button>
      <button formaction="/download-final" formmethod="post">Export to Excel</button>

      <input type="hidden" name="tax_csv"   value="{{ tax_csv }}">
      <input type="hidden" name="exp_csv"   value="{{ exp_csv }}">
      <input type="hidden" name="final_csv" value="{{ final_csv }}">
    </form>

    <div class="chart-container">
  <div class="chart-block">
    <h3>Tax Breakdown Pie</h3>
    <canvas id="taxBreakdownChart" width="200" height="200"></canvas>
  </div>
  <div class="chart-block">
    <h3>Overall Distribution</h3>
    <canvas id="finalChart" width="200" height="200"></canvas>
  </div>
</div>
    <script>
      const origNet = {{ orig_nets|tojson }};

      function recalculate() {
        // Sequentially subtract each expense from its sender’s net
        const running = {...origNet};
        document.querySelectorAll('#expTable tbody tr[data-sender]').forEach(row => {
          const sender = row.dataset.sender;
          const idx    = row.querySelector('input[name^="exp_amt_"]').name.split('_').pop();
          const amt    = parseFloat(row.querySelector(`input[name="exp_amt_${idx}"]`).value) || 0;
          running[sender] -= amt;
          row.querySelector(`input[name="net_after_${idx}"]`).value = running[sender].toFixed(2);
        });

        // Compute overall final net
        let overall = Object.values(running).reduce((sum,v) => sum + v, 0);
        document.getElementById('finalNetCell').innerHTML = `<strong>$${overall.toFixed(2)}</strong>`;
      }

      document.querySelectorAll('#expTable input[name^="exp_amt_"]')
        .forEach(input => input.addEventListener('input', recalculate));
      document.addEventListener('DOMContentLoaded', recalculate);

      function removeRow(btn) { btn.closest('tr').remove(); recalculate(); }

      function addRow() {
        const tbody = document.querySelector('#expTable tbody');
        const idx = tbody.rows.length;
        const row = document.createElement('tr');
        row.setAttribute('data-sender', '');
        row.innerHTML = `
          <td><input name="exp_sender_${idx}" placeholder="Sender"
                    onchange="this.closest('tr').setAttribute('data-sender', this.value)"></td>
          <td><input name="exp_name_${idx}" placeholder="Expense"></td>
          <td><input name="exp_amt_${idx}" oninput="recalculate()"></td>
          <td><input name="net_after_${idx}" readonly></td>
          <td><button type="button" onclick="removeRow(this)">Remove</button></td>
        `;
        tbody.appendChild(row);
        // immediately recalc so new row's net shows up
        recalculate();
      }

    async function saveEntry() {
        // ask the user for a name
        const title = prompt("Enter a name for this entry:");
        if (!title) return;  // user cancelled

        const form = document.getElementById('finalForm');
        const data = new FormData(form);
        data.append('title', title);

        const resp = await fetch('/save-entry', {
          method: 'POST',
          body: data,
          headers: { 'X-Requested-With': 'XMLHttpRequest' }
        });

        if (resp.ok || resp.status===302) {
          alert('Entry "' + title + '" saved successfully!');
          // optionally redirect back to Saved Entries:
          window.location = '/saved-entries';
        } else {
          alert('Save failed.');
        }
      }
      new Chart(
        document.getElementById('taxBreakdownChart'),
        { type:'pie',
          data:{ labels: {{ comp_labels|tojson }},
                 datasets:[{ data: {{ comp_data|tojson }} }]},
          options:{ responsive:false }
        }
      );
      new Chart(
        document.getElementById('finalChart'),
        { type:'pie',
          data:{ labels:['Tax','Expenses','Net'],
                 datasets:[{ data:[{{ total_tax }},{{ total_exp }},{{ total_net }}] }]},
          options:{ responsive:false }
        }
      );
    </script>
  </body>
</html>
'''
statements_html = base_style + nav_html + '''
<html>
  <head><title>Income Statements</title></head>
  <body>
    <h2>Income & Expense Statements by Month</h2>

    <form method="get" style="margin-bottom:1em;">
      <label>Filter by type:
        <select name="type" onchange="this.form.submit()">
          {% for t in types %}
            <option value="{{t}}" {% if t==filter_type %}selected{% endif %}>{{t}}</option>
          {% endfor %}
        </select>
      </label>
    </form>
    <form style="margin-bottom:1em;" action="{{ url_for('download_statements') }}" method="get">
      <input type="hidden" name="type" value="{{ filter_type }}">
      <button type="submit">Export Statements to Excel</button>
    </form>

    {% if not summary %}
      <p><em>No data to show.</em></p>
    {% else %}
      {% for row in summary %}
        <details open style="margin-bottom:2em;">
          <summary style="font-size:1.1em; cursor:pointer;">
            {{ row.month.strftime('%Y-%m') }}
            — Income: ${{ '%.2f'|format(row.inc) }}
            | Expenses: ${{ '%.2f'|format(row.exp) }}
            | Taxes Due: ${{ '%.2f'|format(row.tax) }}
          </summary>
          <div style="padding: 0.5em 1em;">
            <h4>Incomes this month</h4>
            <table>
              <tr>
                <th>Date</th><th>Sender</th><th>Type</th><th>Gross</th><th>Taxes</th><th>Entry</th>
              </tr>
              {% for it in inc_groups[row.month] %}
                <tr>
                  <td>{{ it.date.strftime('%Y-%m-%d') }}</td>
                  <td>{{ it.sender }}</td>
                  <td>{{ it.type }}</td>
                  <td>${{ '%.2f'|format(it.Gross) }}</td>
                  <td>${{ '%.2f'|format(it.taxes_due) }}</td>
                  <td><a href="{{ url_for('view_entry', entry_id=it.entry_id) }}">View</a></td>
                </tr>
              {% endfor %}
            </table>

            <h4 style="margin-top:1em;">Expenses this month</h4>
<table>
  <tr><th>Date</th><th>Sender</th><th>Expense</th><th>Amount</th><th>Entry</th></tr>
  {% for ex in exp_groups[row.month] %}
  <tr>
    <!-- now use ex.date -->
    <td>{{ ex.date.strftime('%Y-%m-%d') }}</td>
    <td>{{ ex.sender }}</td>
    <td>{{ ex.name }}</td>
    <td>${{ '%.2f'|format(ex.amt) }}</td>
    <td><a href="{{ url_for('view_entry', entry_id=ex.entry_id) }}">View</a></td>
  </tr>
  {% endfor %}
</table>
          </div>
        </details>
      {% endfor %}

      <h2>Yearly Summary</h2>
      <table>
        <tr><th>Year</th><th>Total Gross</th><th>Total Expenses</th><th>Total Taxes</th></tr>
        {% for y in yearly %}
          <tr>
            <td>{{ y.year }}</td>
            <td>${{ '%.2f'|format(y.inc_total) }}</td>
            <td>${{ '%.2f'|format(y.exp_total) }}</td>
            <td>${{ '%.2f'|format(y.tax_total) }}</td>
          </tr>
        {% endfor %}
      </table>
    {% endif %}
  </body>
</html>
'''

 
def se_tax(amount: float) -> float:
    return amount * 0.153

def fed_tax(amount: float) -> float:
    return calculate_federal_tax(amount)

def state_tax(amount: float) -> float:
    return amount * louisiana_tax_rate

def show_final_context(tax_csv, exp_csv, final_csv):
    df_tax = pd.read_csv(StringIO(tax_csv))
    comp_labels = ['Self-EE','Fed','State']
    comp_data = [
        df_tax['Self-EE Tax'].sum(),
        df_tax['Fed Tax'].sum(),
        df_tax['State Tax'].sum()
    ]
    exp_lines = []
    total_exp = 0.0
    for i, sender in enumerate(df_tax['Sender']):
        count = int(request.form.get(f'count_{i}', 0))
        for j in range(count):
            name = request.form[f'exp_name_{i}_{j}']
            amt = float(request.form[f'exp_amt_{i}_{j}'])
            exp_lines.append((sender, name, amt))
            total_exp += amt
    total_tax = df_tax['Total Tax'].sum()
    total_net = df_tax['Net'].sum() - total_exp
    orig_nets = df_tax.set_index('Sender')['Net'].to_dict()
    exp_rows = []
    for sender, name, amt in exp_lines:
        total_sender_exp = sum(amt for s, _, amt in exp_lines if s == sender)
        net_profit = round(orig_nets[sender] - total_sender_exp, 2)
        exp_rows.append([sender, name, amt, net_profit])
    tax_rows = [row[1:] for row in df_tax.values.tolist()]
    tax_cols = df_tax.columns.tolist()[1:]
    return dict(
        tax_cols=tax_cols,
        tax_rows=tax_rows,
        exp_rows=exp_rows,
        total_tax=round(total_tax,2),
        total_exp=round(total_exp,2),
        total_net=round(total_net,2),
        comp_labels=comp_labels,
        comp_data=comp_data,
        tax_csv=tax_csv,
        exp_csv=exp_csv,
        final_csv=final_csv,
        orig_nets=orig_nets
    )

@app.route('/', methods=['GET'])
def index():
    return render_template_string(index_html)

@app.route('/tax-entry', methods=['POST'])
def tax_entry():
    n     = int(request.form['num_checks'])
    today = date.today().isoformat()   # "2025-05-08"
    return render_template_string(
      tax_entry_html,
      n=n,
      edit=False,
      senders=[],
      grosses=[],
      today=today
    )


@app.route('/show-taxes', methods=['POST'])
def show_taxes():
    n = int(request.form['num_checks'])
    records = []
    sums = {'se': 0.0, 'fed': 0.0, 'st': 0.0}

    SS_WAGE_BASE = 168_666
    SS_RATE      = 0.124   # Social Security 12.4%
    MED_RATE     = 0.029   # Medicare 2.9%

    for i in range(n):
        sender   = request.form.get(f'sender_{i}', '').strip()
        gross    = float(request.form.get(f'Gross_{i}', '0') or 0)
        inc_type = request.form.get(f'type_{i}', '')
        date_str = request.form.get(f'date_{i}', '')
        check_date = datetime.fromisoformat(date_str).date() if date_str else None

        if inc_type == '1099-NEC':
            med_tax = gross * MED_RATE
            ss_tax  = min(gross, SS_WAGE_BASE) * SS_RATE
            se      = ss_tax + med_tax

            fed = calculate_federal_tax(gross)
            st  = gross * louisiana_tax_rate
        else:
            se = fed = st = 0.0

        total_tax = se + fed + st
        net       = gross - total_tax

        sums['se']  += se
        sums['fed'] += fed
        sums['st']  += st

        records.append([
            sender,
            inc_type,
            check_date,
            round(gross,    2),
            round(se,       2),
            round(fed,      2),
            round(st,       2),
            round(total_tax,2),
            round(net,      2)
        ])

    df = pd.DataFrame(records, columns=[
        'Sender','Type','Date',
        'Gross','Self-EE Tax','Fed Tax','State Tax',
        'Total Tax','Net'
    ])

    return render_template_string(
        show_taxes_html,
        cols        = df.columns.tolist(),
        rows        = df.values.tolist(),
        csv         = df.to_csv(index=False),
        exp_csv     = '',  # start fresh
        comp_labels = ['Self-EE','Fed','State'],
        comp_data   = [
            round(sums['se'],  2),
            round(sums['fed'], 2),
            round(sums['st'],  2)
        ]
    )




@app.route('/expense-entry', methods=['POST'])
def expense_entry():
    tax_csv = request.form['tax_csv']
    exp_csv = request.form.get('exp_csv', '')
    df_tax = pd.read_csv(StringIO(tax_csv))
    num_checks = len(df_tax)
    senders = df_tax['Sender'].tolist()

    saved = {}
    counts = {i: 0 for i in range(num_checks)}
    if exp_csv:
        df_exp = pd.read_csv(StringIO(exp_csv))
        for _, row in df_exp.iterrows():
            sender = row['Sender']
            if sender in senders:
                i = senders.index(sender)
            else:
                continue
            j = counts[i]
            saved[f'exp_name_{i}_{j}'] = row['Expense']
            saved[f'exp_amt_{i}_{j}']  = str(row['Amount'])
            counts[i] += 1

    for i in range(num_checks):
        saved[f'count_{i}'] = counts[i] or 1

    return render_template_string(
        expense_entry_html,
        tax_csv=tax_csv,
        num_checks=num_checks,
        senders=senders,
        saved=saved
    )




@app.route('/show-final', methods=['POST'])
def show_final():
   
    exp_data = {}
    df_tax = pd.read_csv(StringIO(request.form['tax_csv']))
    for i, sender in enumerate(df_tax['Sender']):
        cnt = int(request.form.get(f'count_{i}', 0))
        exp_data[f'count_{i}'] = cnt
        for j in range(cnt):
            exp_data[f'exp_name_{i}_{j}'] = request.form.get(f'exp_name_{i}_{j}', '')
            exp_data[f'exp_amt_{i}_{j}']  = request.form.get(f'exp_amt_{i}_{j}', '')
    session['exp_data'] = exp_data

    tax_csv = request.form['tax_csv']

    df_tax = pd.read_csv(StringIO(tax_csv))

    comp_labels = ['Self-EE', 'Fed', 'State']
    comp_data = [
        df_tax['Self-EE Tax'].sum(),
        df_tax['Fed Tax'].sum(),
        df_tax['State Tax'].sum()
    ]

    exp_lines = []
    total_exp = 0.0

    for i, sender in enumerate(df_tax['Sender']):
      cnt = int(request.form.get(f'count_{i}', 0))
      for j in range(cnt):
        name    = request.form.get(f'exp_name_{i}_{j}', '').strip()
        amt_str = request.form.get(f'exp_amt_{i}_{j}', '').strip()
        if not name or not amt_str:
            continue    # ← skip any row where either field is blank
        amt = float(amt_str)
        exp_lines.append((sender, name, amt))
        total_exp += amt



    total_tax = df_tax['Total Tax'].sum()
    total_net = df_tax['Net'].sum() - total_exp
    orig_nets = df_tax.set_index('Sender')['Net'].to_dict()

    exp_rows = []
    for sender, name, amt in exp_lines:
        total_sender_exp = sum(a for s,_,a in exp_lines if s == sender)
        net_profit = round(orig_nets[sender] - total_sender_exp, 2)
        exp_rows.append([sender, name, amt, net_profit])

    tax_rows = [row[1:] for row in df_tax.values.tolist()]
    tax_cols = df_tax.columns.tolist()[1:]

    return render_template_string(
        final_html,
        tax_cols=tax_cols,
        tax_rows=tax_rows,
        exp_rows=exp_rows,
        total_tax=round(total_tax,2),
        total_exp=round(total_exp,2),
        total_net=round(total_net,2),
        comp_labels=comp_labels,
        comp_data=comp_data,
        tax_csv=tax_csv,
        exp_csv=pd.DataFrame(exp_rows, columns=['Sender','Name','Amount','Net Profit']).to_csv(index=False),
        final_csv=pd.DataFrame([[r[3]] for r in exp_rows], columns=['FinalNet']).to_csv(index=False),
        orig_nets=orig_nets
    )
from io import StringIO

@app.route('/save-entry', methods=['POST'])
def save_entry():
    title = request.form['title'].strip()
    if not title:
        flash("You must supply a name.")
        return redirect(request.referrer or url_for('index'))

    entry = Entry(
        title     = title,
        tax_csv   = request.form['tax_csv'],
        exp_csv   = request.form['exp_csv'],
        final_csv = request.form['final_csv']
    )
    db.session.add(entry)
    db.session.flush()   # give us entry.id without committing

    df_tax = pd.read_csv(StringIO(request.form['tax_csv']), parse_dates=['Date'])
    for _, row in df_tax.iterrows():
        db.session.add( Income(
            entry_id    = entry.id,
            sender      = row['Sender'],
            Gross       = float(row['Gross']),
            income_type = row['Type'],
            date        = row['Date'].date()     # or row['Date'] if it’s already a date
        ) )

    db.session.commit()
    flash(f'Entry "{title}" saved.')
    return redirect(url_for('saved_entries'))



@app.route('/delete-entry/<int:entry_id>', methods=['POST'])
def delete_entry(entry_id):
    Income.query.filter_by(entry_id=entry_id).delete()
    entry = Entry.query.get_or_404(entry_id)
    db.session.delete(entry)
    db.session.commit()
    flash(f"Deleted entry {entry.timestamp:%Y-%m-%d %H:%M:%S}")
    return redirect(url_for('saved_entries'))

@app.route('/saved-entries')
def saved_entries():
    entries = Entry.query.order_by(Entry.timestamp.desc()).all()
    return render_template_string(
        base_style + nav_html + '''
<h2>Saved Entries</h2>
<ul>
  {% for e in entries %}
    <li>
      <strong>{{ e.title }}</strong>
      &nbsp;(<small>{{ e.timestamp.strftime('%Y-%m-%d') }}</small>)
      [<a href="{{ url_for('view_entry', entry_id=e.id) }}">View/Edit</a>]
      [<a href="{{ url_for('download_entry', entry_id=e.id) }}">Download</a>]
      <form action="{{ url_for('delete_entry', entry_id=e.id) }}"
            method="post" style="display:inline;margin-left:8px;">
        <button type="submit">Delete</button>
      </form>
    </li>
  {% endfor %}
</ul>
''', entries=entries
    )


@app.route('/view-entry/<int:entry_id>', methods=['GET'])
def view_entry(entry_id):
    entry = Entry.query.get_or_404(entry_id)

    df_tax = pd.read_csv(StringIO(entry.tax_csv))
    try:
        df_exp = pd.read_csv(StringIO(entry.exp_csv))
    except pd.errors.EmptyDataError:
        df_exp = pd.DataFrame(columns=['Sender','Name','Amount','Net Profit'])

    comp_labels = ['Self-EE','Fed','State']
    comp_data = [
        df_tax['Self-EE Tax'].sum(),
        df_tax['Fed Tax'].sum(),
        df_tax['State Tax'].sum()
    ]
    total_tax = df_tax['Total Tax'].sum()
    total_exp = df_exp['Amount'].sum()
    total_net = df_exp['Net Profit'].sum()
    orig_nets = df_tax.set_index('Sender')['Net'].to_dict()

    tax_cols = df_tax.columns.tolist()[1:]
    tax_rows = [row[1:] for row in df_tax.values.tolist()]
    exp_rows = df_exp[['Sender','Name','Amount','Net Profit']].values.tolist()

    incomes = entry.incomes  # thanks to the backref
    ctx = {
        'tax_cols': tax_cols,
        'tax_rows': tax_rows,
        'exp_rows': exp_rows,
        'comp_labels': comp_labels,
        'comp_data': comp_data,
        'total_tax': round(total_tax,2),
        'total_exp': round(total_exp,2),
        'total_net': round(total_net,2),
        'tax_csv': entry.tax_csv,
        'exp_csv': entry.exp_csv,
        'final_csv': entry.final_csv,
        'orig_nets': orig_nets,
        'view_only': True,
    }
    return render_template_string(final_html, **ctx)

@app.route('/download-final', methods=['POST'])
def download_final():
    df_tax   = pd.read_csv(StringIO(request.form['tax_csv']))
    df_exp   = pd.read_csv(StringIO(request.form['exp_csv']))
    df_final = pd.read_csv(StringIO(request.form['final_csv']))

    total_tax = df_tax['Total Tax'].sum()
    total_exp = df_exp['Amount'].sum()
    total_net = df_final['FinalNet'].sum()

    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'Taxes'
    for row in dataframe_to_rows(df_tax, index=False, header=True):
        ws1.append(row)

    ws2 = wb.create_sheet('Expenses & Net')
    ws2.append(['Sender', 'Expense', 'Amount', 'Net After'])
    for sender, name, amt, net_after in df_exp.values.tolist():
        ws2.append([sender, name, amt, net_after])

    ws3 = wb.create_sheet('Summary')
    ws3.append(['Category', 'Value'])
    ws3.append(['Total Tax', total_tax])
    ws3.append(['Total Expenses', total_exp])
    ws3.append(['Final Net', total_net])

    pie1 = PieChart()
    pie1.title = "Tax Breakdown"
    labels = Reference(ws3, min_col=1, min_row=2, max_row=4)
    data   = Reference(ws3, min_col=2, min_row=2, max_row=4)
    pie1.add_data(data, titles_from_data=False)
    pie1.set_categories(labels)
    ws3.add_chart(pie1, "E2")

    pie2 = PieChart()
    pie2.title = "Overall Distribution"
    pie2.add_data(data, titles_from_data=False)
    pie2.set_categories(labels)
    ws3.add_chart(pie2, "E20")

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        as_attachment=True,
        download_name='report.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/statements')
def statements():
    all_types   = [r[0] for r in Income.query.with_entities(Income.income_type).distinct()]
    types       = ['All'] + all_types
    filter_type = request.args.get('type', 'All')

    q = Income.query
    if filter_type != 'All':
        q = q.filter(Income.income_type == filter_type)
    inc_objs = q.all()

    inc_rows = []
    for inc in inc_objs:
        if inc.income_type == '1099-NEC':
            se   = inc.Gross * 0.153
            fed  = calculate_federal_tax(inc.Gross)
            st   = inc.Gross * louisiana_tax_rate
            tax_amt = round(se + fed + st, 2)
        else:
            tax_amt = 0.0

        inc_rows.append({
            'entry_id':  inc.entry_id,
            'date':      inc.date,
            'sender':    inc.sender,
            'type':      inc.income_type,
            'Gross':     inc.Gross,
            'taxes_due': tax_amt
        })

    inc_df = pd.DataFrame(inc_rows)
    for col, dtype in (('date','datetime64[ns]'),
                       ('Gross','float64'),
                       ('taxes_due','float64')):
        if col not in inc_df:
            inc_df[col] = pd.Series(dtype=dtype)

    inc_df['date']  = pd.to_datetime(inc_df['date'])
    inc_df['month'] = inc_df['date'].dt.to_period('M').dt.to_timestamp()
    inc_df['year']  = inc_df['date'].dt.year

    exp_rows = []
    for entry in Entry.query.all():
        if not entry.exp_csv.strip():
            continue
        df_exp = pd.read_csv(StringIO(entry.exp_csv))
        date_map = {i.sender: i.date for i in entry.incomes}
        for _, r in df_exp.iterrows():
            d = date_map.get(r['Sender'], entry.timestamp.date())
            exp_rows.append({
                'entry_id': entry.id,
                'date':      d,
                'sender':    r['Sender'],
                'name':      r['Name'],
                'amt':       r['Amount']
            })

    exp_df = pd.DataFrame(exp_rows)
    if not exp_df.empty:
        exp_df['date']  = pd.to_datetime(exp_df['date'])
        exp_df['month'] = exp_df['date'].dt.to_period('M').dt.to_timestamp()
        exp_df['year']  = exp_df['date'].dt.year
    else:
        for col in ('date','month','year','amt'):
            exp_df[col] = pd.Series(dtype='datetime64[ns]' if col!='amt' else 'float64')

    months  = sorted(set(inc_df['month']).union(exp_df['month']))
    summary = []
    for m in months:
        summary.append({
            'month': m,
            'inc':   inc_df.loc[inc_df['month']==m, 'Gross'].sum(),
            'exp':   exp_df.loc[exp_df['month']==m, 'amt'].sum(),
            'tax':   inc_df.loc[inc_df['month']==m, 'taxes_due'].sum(),
        })

    inc_groups = defaultdict(list)
    for rec in inc_df.to_dict('records'):
        inc_groups[rec['month']].append(rec)
    exp_groups = defaultdict(list)
    for rec in exp_df.to_dict('records'):
        exp_groups[rec['month']].append(rec)

    yr_inc = inc_df.groupby('year')['Gross'].sum()
    yr_exp = exp_df.groupby('year')['amt'].sum()
    yr_tax = inc_df.groupby('year')['taxes_due'].sum()
    years  = sorted(set(yr_inc.index).union(yr_exp.index).union(yr_tax.index))
    yearly = [{
        'year':      y,
        'inc_total': float(yr_inc.get(y, 0)),
        'exp_total': float(yr_exp.get(y, 0)),
        'tax_total': float(yr_tax.get(y, 0)),
    } for y in years]

    return render_template_string(statements_html,
        types=types,
        filter_type=filter_type,
        summary=summary,
        inc_groups=inc_groups,
        exp_groups=exp_groups,
        yearly=yearly
    )


@app.route('/download-entry/<int:entry_id>')
def download_entry(entry_id):
    e = Entry.query.get_or_404(entry_id)

    df_tax   = pd.read_csv(StringIO(e.tax_csv))
    df_exp   = pd.read_csv(StringIO(e.exp_csv))
    df_final = pd.read_csv(StringIO(e.final_csv))

    total_tax = df_tax['Total Tax'].sum()
    total_exp = df_exp['Amount'].sum()
    total_net = df_final['FinalNet'].sum()

    incs = Income.query.filter_by(entry_id=entry_id).all()

    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'Taxes'
    for row in dataframe_to_rows(df_tax, index=False, header=True):
        ws1.append(row)

    ws2 = wb.create_sheet('Expenses & Net')
    ws2.append(['Sender','Expense','Amount','Net After'])
    for sender, name, amt, net_after in df_exp.values.tolist():
        ws2.append([sender, name, amt, net_after])

    ws3 = wb.create_sheet('Incomes')
    ws3.append(['Date','Sender','Type','Gross','Taxes Due'])
    for inc in incs:
        if inc.income_type == '1099-NEC':
            taxes_due = 0.0
        else:
            fed = calculate_federal_tax(inc.Gross)
            st  = inc.Gross * louisiana_tax_rate
            taxes_due = round(fed + st, 2)
        ws3.append([
            inc.date.strftime('%Y-%m-%d'),
            inc.sender,
            inc.income_type,
            inc.Gross,
            taxes_due
        ])

    ws4 = wb.create_sheet('Summary')
    ws4.append(['Category','Value'])
    ws4.append(['Total Tax',      total_tax])
    ws4.append(['Total Expenses', total_exp])
    ws4.append(['Final Net',      total_net])

    labels = Reference(ws4, min_col=1, min_row=2, max_row=4)
    data   = Reference(ws4, min_col=2, min_row=2, max_row=4)

    pie1 = PieChart(); pie1.title='Tax Breakdown'
    pie1.set_categories(labels); pie1.add_data(data, titles_from_data=False)
    ws4.add_chart(pie1, 'E2')

    pie2 = PieChart(); pie2.title='Overall Distribution'
    pie2.set_categories(labels); pie2.add_data(data, titles_from_data=False)
    ws4.add_chart(pie2, 'E20')

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
      out,
      as_attachment=True,
      download_name=f'report_{entry_id}.xlsx',
      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/download-statements')
def download_statements():
    filter_type = request.args.get('type', 'All')

    q = Income.query
    if filter_type != 'All':
        q = q.filter_by(income_type=filter_type)
    incs = q.order_by(Income.date).all()

    inc_df = pd.DataFrame([{
        'date':         i.date,
        'Gross':        i.Gross,
        'income_type':  i.income_type,
        'entry_id':     i.entry_id
    } for i in incs])
    if inc_df.empty:
        inc_df = pd.DataFrame(columns=['date','Gross','income_type','entry_id'])
    inc_df['date'] = pd.to_datetime(inc_df['date'])

    exp_rows = []
    for e in Entry.query.all():
        try:
            df = pd.read_csv(StringIO(e.exp_csv))
            for _, r in df.iterrows():
                exp_rows.append({
                    'entry_id':   e.id,
                    'entry_date': e.timestamp.date(),
                    'sender':     r['Sender'],
                    'name':       r['Name'],
                    'amt':        r['Amount']
                })
        except:
            pass
    exp_df = pd.DataFrame(exp_rows)
    if not exp_df.empty:
        exp_df['entry_date'] = pd.to_datetime(exp_df['entry_date'])
    else:
        exp_df = pd.DataFrame(columns=['entry_id','entry_date','sender','name','amt'])

    inc_df['month'] = inc_df['date'].dt.to_period('M').dt.to_timestamp()
    inc_df['year']  = inc_df['date'].dt.year

    exp_df['month'] = exp_df['entry_date'].dt.to_period('M').dt.to_timestamp()
    exp_df['year']  = exp_df['entry_date'].dt.year

    mon_inc = inc_df.groupby('month')['Gross'].sum().reset_index(name='inc_total')
    mon_exp = exp_df.groupby('month')['amt'].sum().reset_index(name='exp_total')
    mon_tax = inc_df.assign(
        taxes = inc_df.apply(lambda r: 0.0 if r['income_type']=='1099-NEC'
                             else calculate_federal_tax(r['Gross']) + r['Gross']*louisiana_tax_rate, axis=1)
      ).groupby('month')['taxes'].sum().reset_index(name='tax_total')

    monthly = mon_inc.merge(mon_exp, on='month', how='outer')\
                     .merge(mon_tax, on='month', how='outer')\
                     .fillna(0)

    yr_inc = inc_df.groupby('year')['Gross'].sum().reset_index(name='inc_total')
    yr_exp = exp_df.groupby('year')['amt'].sum().reset_index(name='exp_total')
    yr_tax = inc_df.assign(
        taxes = inc_df.apply(lambda r: 0.0 if r['income_type']=='1099-NEC'
                             else calculate_federal_tax(r['Gross']) + r['Gross']*louisiana_tax_rate, axis=1)
      ).groupby('year')['taxes'].sum().reset_index(name='tax_total')

    yearly = yr_inc.merge(yr_exp, on='year', how='outer')\
                   .merge(yr_tax, on='year', how='outer')\
                   .fillna(0)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'Monthly Summary'
    ws1.append(['Month','Total Income','Total Expenses','Total Taxes Due'])
    for _, r in monthly.sort_values('month').iterrows():
        ws1.append([r['month'].strftime('%Y-%m'), r['inc_total'], r['exp_total'], r['tax_total']])

    ws2 = wb.create_sheet('Yearly Summary')
    ws2.append(['Year','Total Income','Total Expenses','Total Taxes Due'])
    for _, r in yearly.sort_values('year').iterrows():
        ws2.append([int(r['year']), r['inc_total'], r['exp_total'], r['tax_total']])

    ws3 = wb.create_sheet('All Incomes')
    ws3.append(['Date','Sender','Type','Gross','Taxes Due','Entry ID'])
    for _, r in inc_df.sort_values('date').iterrows():
        taxes = 0.0 if r['income_type']=='1099-NEC' else round(
            calculate_federal_tax(r['Gross']) + r['Gross']*louisiana_tax_rate, 2)
        ws3.append([r['date'].strftime('%Y-%m-%d'),
                    '',  # if you want sender you can join via Income model 
                    r['income_type'],
                    r['Gross'],
                    taxes,
                    int(r['entry_id'])])

    ws4 = wb.create_sheet('All Expenses')
    ws4.append(['Entry Date','Sender','Expense','Amount','Entry ID'])
    for _, r in exp_df.sort_values('entry_date').iterrows():
        ws4.append([r['entry_date'].strftime('%Y-%m-%d'),
                    r['sender'],
                    r['name'],
                    r['amt'],
                    int(r['entry_id'])])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
      out,
      as_attachment=True,
      download_name='statements.xlsx',
      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
if __name__ == '__main__':
   
     app.run(debug=True)
     


@app.route('/expense-entry', methods=['POST'])
def expense_entry():
    tax_csv = request.form['tax_csv']
    exp_csv = request.form.get('exp_csv', '')
    df_tax = pd.read_csv(StringIO(tax_csv))
    num_checks = len(df_tax)
    senders = df_tax['Sender'].tolist()

    saved = {}
    counts = {i: 0 for i in range(num_checks)}
    if exp_csv:
        df_exp = pd.read_csv(StringIO(exp_csv))
        for _, row in df_exp.iterrows():
            sender = row['Sender']
            if sender in senders:
                i = senders.index(sender)
            else:
                continue
            j = counts[i]
            saved[f'exp_name_{i}_{j}'] = row['Expense']
            saved[f'exp_amt_{i}_{j}']  = str(row['Amount'])
            counts[i] += 1

    for i in range(num_checks):
        saved[f'count_{i}'] = counts[i] or 1

    return render_template_string(
        expense_entry_html,
        tax_csv=tax_csv,
        num_checks=num_checks,
        senders=senders,
        saved=saved
    )




@app.route('/show-final', methods=['POST'])
def show_final():
   
    exp_data = {}
    df_tax = pd.read_csv(StringIO(request.form['tax_csv']))
    for i, sender in enumerate(df_tax['Sender']):
        cnt = int(request.form.get(f'count_{i}', 0))
        exp_data[f'count_{i}'] = cnt
        for j in range(cnt):
            exp_data[f'exp_name_{i}_{j}'] = request.form.get(f'exp_name_{i}_{j}', '')
            exp_data[f'exp_amt_{i}_{j}']  = request.form.get(f'exp_amt_{i}_{j}', '')
    session['exp_data'] = exp_data

    tax_csv = request.form['tax_csv']

    df_tax = pd.read_csv(StringIO(tax_csv))

    comp_labels = ['Self-EE', 'Fed', 'State']
    comp_data = [
        df_tax['Self-EE Tax'].sum(),
        df_tax['Fed Tax'].sum(),
        df_tax['State Tax'].sum()
    ]

    exp_lines = []
    total_exp  = 0.0

    for i, sender in enumerate(df_tax['Sender']):
        cnt = int(request.form.get(f'count_{i}', 0))
        for j in range(cnt):
            name    = request.form.get(f'exp_name_{i}_{j}', '').strip()
            amt_str = request.form.get(f'exp_amt_{i}_{j}', '').strip()
            if not name or not amt_str:
                continue
            amt = float(amt_str)
            exp_lines.append((sender, name, amt))
            total_exp += amt



    total_tax = df_tax['Total Tax'].sum()
    total_net = df_tax['Net'].sum() - total_exp
    orig_nets = df_tax.set_index('Sender')['Net'].to_dict()

    exp_rows = []
    for sender, name, amt in exp_lines:
        total_sender_exp = sum(a for s,_,a in exp_lines if s == sender)
        net_profit = round(orig_nets[sender] - total_sender_exp, 2)
        exp_rows.append([sender, name, amt, net_profit])

    tax_rows = [row[1:] for row in df_tax.values.tolist()]
    tax_cols = df_tax.columns.tolist()[1:]

    return render_template_string(
        final_html,
        tax_cols=tax_cols,
        tax_rows=tax_rows,
        exp_rows=exp_rows,
        total_tax=round(total_tax,2),
        total_exp=round(total_exp,2),
        total_net=round(total_net,2),
        comp_labels=comp_labels,
        comp_data=comp_data,
        tax_csv=tax_csv,
        exp_csv=pd.DataFrame(exp_rows, columns=['Sender','Name','Amount','Net Profit']).to_csv(index=False),
        final_csv=pd.DataFrame([[r[3]] for r in exp_rows], columns=['FinalNet']).to_csv(index=False),
        orig_nets=orig_nets
    )
from io import StringIO

@app.route('/save-entry', methods=['POST'])
def save_entry():
    title = request.form['title'].strip()
    if not title:
        flash("You must supply a name.")
        return redirect(request.referrer or url_for('index'))

    entry = Entry(
        title     = title,
        tax_csv   = request.form['tax_csv'],
        exp_csv   = request.form['exp_csv'],
        final_csv = request.form['final_csv']
    )
    db.session.add(entry)
    db.session.flush()   # give us entry.id without committing

    df_tax = pd.read_csv(StringIO(request.form['tax_csv']), parse_dates=['Date'])
    for _, row in df_tax.iterrows():
        db.session.add( Income(
            entry_id    = entry.id,
            sender      = row['Sender'],
            Gross       = float(row['Gross']),
            income_type = row['Type'],
            date        = row['Date'].date()     # or row['Date'] if it’s already a date
        ) )

    db.session.commit()
    flash(f'Entry "{title}" saved.')
    return redirect(url_for('saved_entries'))



@app.route('/delete-entry/<int:entry_id>', methods=['POST'])
def delete_entry(entry_id):
    Income.query.filter_by(entry_id=entry_id).delete()
    entry = Entry.query.get_or_404(entry_id)
    db.session.delete(entry)
    db.session.commit()
    flash(f"Deleted entry {entry.timestamp:%Y-%m-%d %H:%M:%S}")
    return redirect(url_for('saved_entries'))

@app.route('/saved-entries')
def saved_entries():
    entries = Entry.query.order_by(Entry.timestamp.desc()).all()
    return render_template_string(
        base_style + nav_html + '''
<h2>Saved Entries</h2>
<ul>
  {% for e in entries %}
    <li>
      <strong>{{ e.title }}</strong>
      &nbsp;(<small>{{ e.timestamp.strftime('%Y-%m-%d') }}</small>)
      [<a href="{{ url_for('view_entry', entry_id=e.id) }}">View/Edit</a>]
      [<a href="{{ url_for('download_entry', entry_id=e.id) }}">Download</a>]
      <form action="{{ url_for('delete_entry', entry_id=e.id) }}"
            method="post" style="display:inline;margin-left:8px;">
        <button type="submit">Delete</button>
      </form>
    </li>
  {% endfor %}
</ul>
''', entries=entries
    )


@app.route('/view-entry/<int:entry_id>', methods=['GET'])
def view_entry(entry_id):
    entry = Entry.query.get_or_404(entry_id)

    df_tax = pd.read_csv(StringIO(entry.tax_csv))
    try:
        df_exp = pd.read_csv(StringIO(entry.exp_csv))
    except pd.errors.EmptyDataError:
        df_exp = pd.DataFrame(columns=['Sender','Name','Amount','Net Profit'])

    comp_labels = ['Self-EE','Fed','State']
    comp_data = [
        df_tax['Self-EE Tax'].sum(),
        df_tax['Fed Tax'].sum(),
        df_tax['State Tax'].sum()
    ]
    total_tax = df_tax['Total Tax'].sum()
    total_exp = df_exp['Amount'].sum()
    total_net = df_exp['Net Profit'].sum()
    orig_nets = df_tax.set_index('Sender')['Net'].to_dict()

    tax_cols = df_tax.columns.tolist()[1:]
    tax_rows = [row[1:] for row in df_tax.values.tolist()]
    exp_rows = df_exp[['Sender','Name','Amount','Net Profit']].values.tolist()

    incomes = entry.incomes  # thanks to the backref
    ctx = {
        'tax_cols': tax_cols,
        'tax_rows': tax_rows,
        'exp_rows': exp_rows,
        'comp_labels': comp_labels,
        'comp_data': comp_data,
        'total_tax': round(total_tax,2),
        'total_exp': round(total_exp,2),
        'total_net': round(total_net,2),
        'tax_csv': entry.tax_csv,
        'exp_csv': entry.exp_csv,
        'final_csv': entry.final_csv,
        'orig_nets': orig_nets,
        'view_only': True,
    }
    return render_template_string(final_html, **ctx)

@app.route('/download-final', methods=['POST'])
def download_final():
    df_tax   = pd.read_csv(StringIO(request.form['tax_csv']))
    df_exp   = pd.read_csv(StringIO(request.form['exp_csv']))
    df_final = pd.read_csv(StringIO(request.form['final_csv']))

    total_tax = df_tax['Total Tax'].sum()
    total_exp = df_exp['Amount'].sum()
    total_net = df_final['FinalNet'].sum()

    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'Taxes'
    for row in dataframe_to_rows(df_tax, index=False, header=True):
        ws1.append(row)

    ws2 = wb.create_sheet('Expenses & Net')
    ws2.append(['Sender', 'Expense', 'Amount', 'Net After'])
    for sender, name, amt, net_after in df_exp.values.tolist():
        ws2.append([sender, name, amt, net_after])

    ws3 = wb.create_sheet('Summary')
    ws3.append(['Category', 'Value'])
    ws3.append(['Total Tax', total_tax])
    ws3.append(['Total Expenses', total_exp])
    ws3.append(['Final Net', total_net])

    pie1 = PieChart()
    pie1.title = "Tax Breakdown"
    labels = Reference(ws3, min_col=1, min_row=2, max_row=4)
    data   = Reference(ws3, min_col=2, min_row=2, max_row=4)
    pie1.add_data(data, titles_from_data=False)
    pie1.set_categories(labels)
    ws3.add_chart(pie1, "E2")

    pie2 = PieChart()
    pie2.title = "Overall Distribution"
    pie2.add_data(data, titles_from_data=False)
    pie2.set_categories(labels)
    ws3.add_chart(pie2, "E20")

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        as_attachment=True,
        download_name='report.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

from collections import defaultdict
from datetime import datetime

@app.route('/statements')
def statements():
    all_types   = [r[0] for r in Income.query.with_entities(Income.income_type).distinct()]
    types       = ['All'] + all_types
    filter_type = request.args.get('type', 'All')

    q = Income.query
    if filter_type != 'All':
        q = q.filter(Income.income_type == filter_type)
    inc_objs = q.all()

    inc_rows = []
    for inc in inc_objs:
        if inc.income_type == '1099-NEC':
            se   = inc.Gross * 0.153
            fed  = calculate_federal_tax(inc.Gross)
            st   = inc.Gross * louisiana_tax_rate
            tax_amt = round(se + fed + st, 2)
        else:
            tax_amt = 0.0

        inc_rows.append({
            'entry_id':  inc.entry_id,
            'date':      inc.date,
            'sender':    inc.sender,
            'type':      inc.income_type,
            'Gross':     inc.Gross,
            'taxes_due': tax_amt
        })

    inc_df = pd.DataFrame(inc_rows)
    for col, dtype in (('date','datetime64[ns]'),
                       ('Gross','float64'),
                       ('taxes_due','float64')):
        if col not in inc_df:
            inc_df[col] = pd.Series(dtype=dtype)

    inc_df['date']  = pd.to_datetime(inc_df['date'])
    inc_df['month'] = inc_df['date'].dt.to_period('M').dt.to_timestamp()
    inc_df['year']  = inc_df['date'].dt.year

    exp_rows = []
    for entry in Entry.query.all():
        if not entry.exp_csv.strip():
            continue
        df_exp = pd.read_csv(StringIO(entry.exp_csv))
        date_map = {i.sender: i.date for i in entry.incomes}
        for _, r in df_exp.iterrows():
            d = date_map.get(r['Sender'], entry.timestamp.date())
            exp_rows.append({
                'entry_id': entry.id,
                'date':      d,
                'sender':    r['Sender'],
                'name':      r['Name'],
                'amt':       r['Amount']
            })

    exp_df = pd.DataFrame(exp_rows)
    if not exp_df.empty:
        exp_df['date']  = pd.to_datetime(exp_df['date'])
        exp_df['month'] = exp_df['date'].dt.to_period('M').dt.to_timestamp()
        exp_df['year']  = exp_df['date'].dt.year
    else:
        for col in ('date','month','year','amt'):
            exp_df[col] = pd.Series(dtype='datetime64[ns]' if col!='amt' else 'float64')

    months  = sorted(set(inc_df['month']).union(exp_df['month']))
    summary = []
    for m in months:
        summary.append({
            'month': m,
            'inc':   inc_df.loc[inc_df['month']==m, 'Gross'].sum(),
            'exp':   exp_df.loc[exp_df['month']==m, 'amt'].sum(),
            'tax':   inc_df.loc[inc_df['month']==m, 'taxes_due'].sum(),
        })

    inc_groups = defaultdict(list)
    for rec in inc_df.to_dict('records'):
        inc_groups[rec['month']].append(rec)
    exp_groups = defaultdict(list)
    for rec in exp_df.to_dict('records'):
        exp_groups[rec['month']].append(rec)

    yr_inc = inc_df.groupby('year')['Gross'].sum()
    yr_exp = exp_df.groupby('year')['amt'].sum()
    yr_tax = inc_df.groupby('year')['taxes_due'].sum()
    years  = sorted(set(yr_inc.index).union(yr_exp.index).union(yr_tax.index))
    yearly = [{
        'year':      y,
        'inc_total': float(yr_inc.get(y, 0)),
        'exp_total': float(yr_exp.get(y, 0)),
        'tax_total': float(yr_tax.get(y, 0)),
    } for y in years]

    return render_template_string(statements_html,
        types=types,
        filter_type=filter_type,
        summary=summary,
        inc_groups=inc_groups,
        exp_groups=exp_groups,
        yearly=yearly
    )



from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference

@app.route('/download-entry/<int:entry_id>')
def download_entry(entry_id):
    e = Entry.query.get_or_404(entry_id)

    df_tax   = pd.read_csv(StringIO(e.tax_csv))
    df_exp   = pd.read_csv(StringIO(e.exp_csv))
    df_final = pd.read_csv(StringIO(e.final_csv))

    total_tax = df_tax['Total Tax'].sum()
    total_exp = df_exp['Amount'].sum()
    total_net = df_final['FinalNet'].sum()

    incs = Income.query.filter_by(entry_id=entry_id).all()

    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'Taxes'
    for row in dataframe_to_rows(df_tax, index=False, header=True):
        ws1.append(row)

    ws2 = wb.create_sheet('Expenses & Net')
    ws2.append(['Sender','Expense','Amount','Net After'])
    for sender, name, amt, net_after in df_exp.values.tolist():
        ws2.append([sender, name, amt, net_after])

    ws3 = wb.create_sheet('Incomes')
    ws3.append(['Date','Sender','Type','Gross','Taxes Due'])
    for inc in incs:
        if inc.income_type == '1099-NEC':
            taxes_due = 0.0
        else:
            fed = calculate_federal_tax(inc.Gross)
            st  = inc.Gross * louisiana_tax_rate
            taxes_due = round(fed + st, 2)
        ws3.append([
            inc.date.strftime('%Y-%m-%d'),
            inc.sender,
            inc.income_type,
            inc.Gross,
            taxes_due
        ])

    ws4 = wb.create_sheet('Summary')
    ws4.append(['Category','Value'])
    ws4.append(['Total Tax',      total_tax])
    ws4.append(['Total Expenses', total_exp])
    ws4.append(['Final Net',      total_net])

    labels = Reference(ws4, min_col=1, min_row=2, max_row=4)
    data   = Reference(ws4, min_col=2, min_row=2, max_row=4)

    pie1 = PieChart(); pie1.title='Tax Breakdown'
    pie1.set_categories(labels); pie1.add_data(data, titles_from_data=False)
    ws4.add_chart(pie1, 'E2')

    pie2 = PieChart(); pie2.title='Overall Distribution'
    pie2.set_categories(labels); pie2.add_data(data, titles_from_data=False)
    ws4.add_chart(pie2, 'E20')

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
      out,
      as_attachment=True,
      download_name=f'report_{entry_id}.xlsx',
      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/download-statements')
def download_statements():
    filter_type = request.args.get('type', 'All')

    q = Income.query
    if filter_type != 'All':
        q = q.filter_by(income_type=filter_type)
    incs = q.order_by(Income.date).all()

    inc_df = pd.DataFrame([{
        'date':         i.date,
        'Gross':        i.Gross,
        'income_type':  i.income_type,
        'entry_id':     i.entry_id
    } for i in incs])
    if inc_df.empty:
        inc_df = pd.DataFrame(columns=['date','Gross','income_type','entry_id'])
    inc_df['date'] = pd.to_datetime(inc_df['date'])

    exp_rows = []
    for e in Entry.query.all():
        try:
            df = pd.read_csv(StringIO(e.exp_csv))
            for _, r in df.iterrows():
                exp_rows.append({
                    'entry_id':   e.id,
                    'entry_date': e.timestamp.date(),
                    'sender':     r['Sender'],
                    'name':       r['Name'],
                    'amt':        r['Amount']
                })
        except:
            pass
    exp_df = pd.DataFrame(exp_rows)
    if not exp_df.empty:
        exp_df['entry_date'] = pd.to_datetime(exp_df['entry_date'])
    else:
        exp_df = pd.DataFrame(columns=['entry_id','entry_date','sender','name','amt'])

    inc_df['month'] = inc_df['date'].dt.to_period('M').dt.to_timestamp()
    inc_df['year']  = inc_df['date'].dt.year

    exp_df['month'] = exp_df['entry_date'].dt.to_period('M').dt.to_timestamp()
    exp_df['year']  = exp_df['entry_date'].dt.year

    mon_inc = inc_df.groupby('month')['Gross'].sum().reset_index(name='inc_total')
    mon_exp = exp_df.groupby('month')['amt'].sum().reset_index(name='exp_total')
    mon_tax = inc_df.assign(
        taxes = inc_df.apply(lambda r: 0.0 if r['income_type']=='1099-NEC'
                             else calculate_federal_tax(r['Gross']) + r['Gross']*louisiana_tax_rate, axis=1)
      ).groupby('month')['taxes'].sum().reset_index(name='tax_total')

    monthly = mon_inc.merge(mon_exp, on='month', how='outer')\
                     .merge(mon_tax, on='month', how='outer')\
                     .fillna(0)

    yr_inc = inc_df.groupby('year')['Gross'].sum().reset_index(name='inc_total')
    yr_exp = exp_df.groupby('year')['amt'].sum().reset_index(name='exp_total')
    yr_tax = inc_df.assign(
        taxes = inc_df.apply(lambda r: 0.0 if r['income_type']=='1099-NEC'
                             else calculate_federal_tax(r['Gross']) + r['Gross']*louisiana_tax_rate, axis=1)
      ).groupby('year')['taxes'].sum().reset_index(name='tax_total')

    yearly = yr_inc.merge(yr_exp, on='year', how='outer')\
                   .merge(yr_tax, on='year', how='outer')\
                   .fillna(0)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'Monthly Summary'
    ws1.append(['Month','Total Income','Total Expenses','Total Taxes Due'])
    for _, r in monthly.sort_values('month').iterrows():
        ws1.append([r['month'].strftime('%Y-%m'), r['inc_total'], r['exp_total'], r['tax_total']])

    ws2 = wb.create_sheet('Yearly Summary')
    ws2.append(['Year','Total Income','Total Expenses','Total Taxes Due'])
    for _, r in yearly.sort_values('year').iterrows():
        ws2.append([int(r['year']), r['inc_total'], r['exp_total'], r['tax_total']])

    ws3 = wb.create_sheet('All Incomes')
    ws3.append(['Date','Sender','Type','Gross','Taxes Due','Entry ID'])
    for _, r in inc_df.sort_values('date').iterrows():
        taxes = 0.0 if r['income_type']=='1099-NEC' else round(
            calculate_federal_tax(r['Gross']) + r['Gross']*louisiana_tax_rate, 2)
        ws3.append([r['date'].strftime('%Y-%m-%d'),
                    '',  # if you want sender you can join via Income model 
                    r['income_type'],
                    r['Gross'],
                    taxes,
                    int(r['entry_id'])])

    ws4 = wb.create_sheet('All Expenses')
    ws4.append(['Entry Date','Sender','Expense','Amount','Entry ID'])
    for _, r in exp_df.sort_values('entry_date').iterrows():
        ws4.append([r['entry_date'].strftime('%Y-%m-%d'),
                    r['sender'],
                    r['name'],
                    r['amt'],
                    int(r['entry_id'])])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
      out,
      as_attachment=True,
      download_name='statements.xlsx',
      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
if __name__ == '__main__':
   
     app.run(debug=True)
     
